"""
Import ring specs from McCulloch_Ring_Data_Extracted.xlsx into the database.
Matches products by SKU (BJ-001, BJ-002, etc.)
Run: python Server/scripts/import-ring-specs.py
"""

import re
import sys
import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from urllib.parse import urlparse, unquote
from pathlib import Path
import os

# ── Config ────────────────────────────────────────────────────────────────────
DB_URL   = os.environ.get('DATABASE_URL', 'postgresql://mcculloch_admin:#mcculloch_admin#20026@31.97.116.89:5433/mcculloch_db')
XLSX     = Path(__file__).parent.parent.parent / 'McCulloch_Ring_Data_Extracted.xlsx'

# ── Helpers ───────────────────────────────────────────────────────────────────
def parse_ct(val):
    """'0.65 Ct' or '1.00 Ct' → 0.65"""
    if not val:
        return None
    m = re.search(r'(\d+\.?\d*)', str(val))
    return float(m.group(1)) if m else None

def parse_pcs(val):
    """'1 pcs' or '2 pcs' → 1"""
    if not val:
        return None
    m = re.search(r'(\d+)', str(val))
    return int(m.group(1)) if m else None

def parse_float(val):
    if val is None:
        return None
    try:
        return float(str(val).strip())
    except Exception:
        return None

def is_vlookup(val):
    return val and str(val).startswith('=')

# ── Load Excel ────────────────────────────────────────────────────────────────
print(f'Loading: {XLSX}')
wb = openpyxl.load_workbook(str(XLSX))

# Sheet 1: Ring Specs (main data)
ws_main = wb['Ring Specs']
# Sheet 2: Side Stone Breakdown
ws_side = wb['Side Stone Breakdown']

# Read main specs rows
main_rows = []
headers = None
for i, row in enumerate(ws_main.iter_rows(values_only=True)):
    if i == 0:
        headers = [str(c).strip() if c else '' for c in row]
        continue
    if not row[0]:
        continue
    rec = dict(zip(headers, row))
    main_rows.append(rec)

print(f'Main rows: {len(main_rows)}')

# Read side stone rows
side_rows = []
sh = None
for i, row in enumerate(ws_side.iter_rows(values_only=True)):
    if i == 0:
        sh = [str(c).strip() if c else '' for c in row]
        continue
    if not row[0]:
        continue
    rec = dict(zip(sh, row))
    side_rows.append(rec)

print(f'Side stone rows: {len(side_rows)}')

# ── Connect DB ────────────────────────────────────────────────────────────────
u = urlparse(DB_URL)
conn = psycopg2.connect(
    host=u.hostname, port=u.port or 5432,
    dbname=u.path.lstrip('/'),
    user=u.username, password=unquote(u.password),
    sslmode='require'
)
cur = conn.cursor()
print('Connected to DB')

# Fetch all products → {sku: product_id}
cur.execute("SELECT id, sku FROM products WHERE sku IS NOT NULL")
sku_map = {row[1].strip(): row[0] for row in cur.fetchall()}
print(f'Products in DB: {len(sku_map)}')

# ── Import ring specs ─────────────────────────────────────────────────────────
specs_inserted = 0
specs_skipped  = 0
no_match       = []

for rec in main_rows:
    bj_code = str(rec.get('BJ Code', '') or '').strip()
    # Skip letter-variant rows (BJ-101(A) etc.) — those are sub-variants
    if re.search(r'\([A-Fa-f]\)', bj_code):
        specs_skipped += 1
        continue

    product_id = sku_map.get(bj_code)
    if not product_id:
        no_match.append(bj_code)
        continue

    silver_wt    = parse_float(rec.get('Silver Wt (g)'))
    gold_9kt_wt  = parse_float(rec.get('9kt Wt (g)'))
    gold_14kt_wt = parse_float(rec.get('14kt Wt (g)'))
    gold_18kt_wt = parse_float(rec.get('18kt Wt (g)'))
    platinum_wt  = parse_float(rec.get('Platinum Wt (g)'))

    cs1_shape_raw = rec.get('CS1 Shape', '')
    cs1_shape  = None if is_vlookup(cs1_shape_raw) else (str(cs1_shape_raw).strip() or None)
    cs1_size   = str(rec.get('CS1 Size', '') or '').strip() or None
    cs1_carats = parse_ct(rec.get('CS1 Ct'))
    cs1_pieces = parse_pcs(rec.get('CS1 Pcs'))

    cs2_shape_raw = rec.get('CS2 Shape', '')
    cs2_shape  = None if is_vlookup(cs2_shape_raw) else (str(cs2_shape_raw).strip() or None)
    cs2_size   = str(rec.get('CS2 Size', '') or '').strip() or None
    cs2_carats = parse_ct(rec.get('CS2 Ct'))
    cs2_pieces = parse_pcs(rec.get('CS2 Pcs'))

    stone_shape  = str(rec.get('Stone Shape', '') or '').strip() or None
    stone_type   = str(rec.get('Stone Type', '') or '').strip() or None
    ring_styles  = str(rec.get('Ring Styles', '') or '').strip() or None

    cur.execute("""
        INSERT INTO product_ring_specs
          (id, product_id, silver_wt, gold_9kt_wt, gold_14kt_wt, gold_18kt_wt, platinum_wt,
           cs1_shape, cs1_size, cs1_carats, cs1_pieces,
           cs2_shape, cs2_size, cs2_carats, cs2_pieces,
           stone_shape, stone_type, ring_styles, created_at, updated_at)
        VALUES
          (gen_random_uuid(), %s, %s, %s, %s, %s, %s,
           %s, %s, %s, %s,
           %s, %s, %s, %s,
           %s, %s, %s, NOW(), NOW())
        ON CONFLICT (product_id) DO UPDATE SET
          silver_wt    = EXCLUDED.silver_wt,
          gold_9kt_wt  = EXCLUDED.gold_9kt_wt,
          gold_14kt_wt = EXCLUDED.gold_14kt_wt,
          gold_18kt_wt = EXCLUDED.gold_18kt_wt,
          platinum_wt  = EXCLUDED.platinum_wt,
          cs1_shape    = EXCLUDED.cs1_shape,
          cs1_size     = EXCLUDED.cs1_size,
          cs1_carats   = EXCLUDED.cs1_carats,
          cs1_pieces   = EXCLUDED.cs1_pieces,
          cs2_shape    = EXCLUDED.cs2_shape,
          cs2_size     = EXCLUDED.cs2_size,
          cs2_carats   = EXCLUDED.cs2_carats,
          cs2_pieces   = EXCLUDED.cs2_pieces,
          stone_shape  = EXCLUDED.stone_shape,
          stone_type   = EXCLUDED.stone_type,
          ring_styles  = EXCLUDED.ring_styles,
          updated_at   = NOW()
    """, (
        product_id, silver_wt, gold_9kt_wt, gold_14kt_wt, gold_18kt_wt, platinum_wt,
        cs1_shape, cs1_size, cs1_carats, cs1_pieces,
        cs2_shape, cs2_size, cs2_carats, cs2_pieces,
        stone_shape, stone_type, ring_styles
    ))
    specs_inserted += 1

# ── Import side stones ────────────────────────────────────────────────────────
side_inserted = 0

# Group by BJ code
from collections import defaultdict
by_bj = defaultdict(list)
for row in side_rows:
    bj = str(row.get('BJ Code', '') or '').strip()
    by_bj[bj].append(row)

for bj_code, stones in by_bj.items():
    product_id = sku_map.get(bj_code)
    if not product_id:
        continue

    # Delete existing side stones for this product
    cur.execute("DELETE FROM product_side_stones WHERE product_id = %s", (product_id,))

    for i, stone in enumerate(stones):
        shape      = str(stone.get('Shape', '') or '').strip() or None
        dimensions = str(stone.get('Dimensions (mm)', '') or '').strip() or None
        pieces_raw = stone.get('Number of Pieces')
        pieces     = int(pieces_raw) if pieces_raw and str(pieces_raw).strip().isdigit() else None
        carats_raw = stone.get('Carat Weight')
        carats     = parse_float(carats_raw)
        raw_entry  = str(stone.get('Raw Entry', '') or '').strip() or None

        cur.execute("""
            INSERT INTO product_side_stones
              (id, product_id, shape, dimensions, pieces, carats, raw_entry, sort_order, created_at, updated_at)
            VALUES
              (gen_random_uuid(), %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
        """, (product_id, shape, dimensions, pieces, carats, raw_entry, i))
        side_inserted += 1

conn.commit()
cur.close()
conn.close()

print(f'\n✅ Done!')
print(f'   Ring specs upserted : {specs_inserted}')
print(f'   Skipped (variants)  : {specs_skipped}')
print(f'   Side stones inserted: {side_inserted}')
print(f'   No DB match ({len(no_match)}): {no_match[:20]}')

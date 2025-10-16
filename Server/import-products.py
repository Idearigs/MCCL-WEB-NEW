"""
Product Import Script
Imports engagement ring products from Excel file into the database
"""

import openpyxl
import sys
import os
import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Database connection
DB_CONFIG = {
    'host': os.getenv('PG_HOST', 'localhost'),
    'port': os.getenv('PG_PORT', 5432),
    'database': os.getenv('PG_DATABASE'),
    'user': os.getenv('PG_USERNAME'),
    'password': os.getenv('PG_PASSWORD')
}

def get_db_connection():
    """Create database connection"""
    return psycopg2.connect(**DB_CONFIG)

def get_category_id(cursor, category_name='Rings'):
    """Get category ID by name"""
    cursor.execute("SELECT id FROM categories WHERE name = %s", (category_name,))
    result = cursor.fetchone()
    return result['id'] if result else None

def get_jewelry_sub_type_id(cursor, sub_type_name='Engagement Rings'):
    """Get jewelry sub type ID"""
    cursor.execute("SELECT id FROM jewelry_sub_types WHERE name = %s", (sub_type_name,))
    result = cursor.fetchone()
    return result['id'] if result else None

def get_ring_style_mapping(cursor):
    """Get all ring styles and create name->id mapping"""
    cursor.execute("SELECT id, name FROM ring_types")
    return {row['name']: row['id'] for row in cursor.fetchall()}

def get_stone_shape_mapping(cursor):
    """Get all stone shapes and create name->id mapping"""
    cursor.execute("SELECT id, name FROM stone_shapes")
    return {row['name']: row['id'] for row in cursor.fetchall()}

def get_stone_type_mapping(cursor):
    """Get all stone types and create name->id mapping"""
    cursor.execute("SELECT id, name FROM stone_types")
    return {row['name']: row['id'] for row in cursor.fetchall()}

def product_exists(cursor, sku):
    """Check if product with SKU already exists"""
    cursor.execute("SELECT id FROM products WHERE sku = %s", (sku,))
    return cursor.fetchone() is not None

def generate_slug(name):
    """Generate URL-friendly slug from product name"""
    return name.lower().replace(' ', '-').replace("'", '')

def import_products():
    """Main import function"""
    excel_file = r'C:\Users\minuk\Videos\Diamond Ring Set_Aravinth.xlsx'

    print('='*80)
    print('PRODUCT IMPORT SCRIPT')
    print('='*80)
    print()

    # Load Excel file
    print('Loading Excel file...')
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    # Find Engagement Ring sheet
    engagement_sheet = None
    for sheet_name in wb.sheetnames:
        if 'engagement' in sheet_name.lower():
            engagement_sheet = wb[sheet_name]
            print(f'Found sheet: {sheet_name}')
            break

    if not engagement_sheet:
        print('ERROR: Engagement Ring sheet not found!')
        return

    print(f'Sheet has {engagement_sheet.max_row - 1} data rows')
    print()

    # Connect to database
    print('Connecting to database...')
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    try:
        # Get mappings
        print('Loading database mappings...')
        category_id = get_category_id(cursor)
        jewelry_sub_type_id = get_jewelry_sub_type_id(cursor)
        ring_style_map = get_ring_style_mapping(cursor)
        stone_shape_map = get_stone_shape_mapping(cursor)
        stone_type_map = get_stone_type_mapping(cursor)

        if not category_id:
            print('ERROR: Rings category not found!')
            return

        if not jewelry_sub_type_id:
            print('ERROR: Engagement Rings sub-type not found!')
            return

        print(f'  - Category ID: {category_id}')
        print(f'  - Jewelry Sub Type ID: {jewelry_sub_type_id}')
        print(f'  - Ring Styles: {len(ring_style_map)} loaded')
        print(f'  - Stone Shapes: {len(stone_shape_map)} loaded')
        print(f'  - Stone Types: {len(stone_type_map)} loaded')
        print()

        # Process each row
        imported_count = 0
        skipped_count = 0
        error_count = 0

        for row_idx in range(2, engagement_sheet.max_row + 1):
            row = engagement_sheet[row_idx]

            # Read row data
            sku = str(row[0].value).strip() if row[0].value else None
            name = str(row[2].value).strip() if row[2].value else None
            description = str(row[3].value).strip() if row[3].value else None
            ring_style_1 = str(row[4].value).strip() if row[4].value else None
            ring_style_2 = str(row[5].value).strip() if row[5].value else None
            ring_style_3 = str(row[6].value).strip() if row[6].value else None
            ring_style_4 = str(row[7].value).strip() if row[7].value else None
            ring_style_5 = str(row[8].value).strip() if row[8].value else None
            stone_shape = str(row[9].value).strip() if row[9].value else None
            stone_type = str(row[10].value).strip() if row[10].value else None

            # Skip rows without SKU or name
            if not sku or not name or name == '#N/A':
                skipped_count += 1
                continue

            # Check if product already exists
            if product_exists(cursor, sku):
                print(f'  [{row_idx-1}/{engagement_sheet.max_row-1}] SKIPPED: {sku} - {name} (already exists)')
                skipped_count += 1
                continue

            try:
                # Map ring styles to IDs
                ring_style_ids = []
                for style_name in [ring_style_1, ring_style_2, ring_style_3, ring_style_4, ring_style_5]:
                    if style_name and style_name in ring_style_map:
                        ring_style_ids.append(ring_style_map[style_name])
                    else:
                        ring_style_ids.append(None)

                # Map stone shape to ID
                stone_shape_id = stone_shape_map.get(stone_shape) if stone_shape else None

                # Map stone type to ID
                stone_type_id = stone_type_map.get(stone_type) if stone_type else None

                # Generate slug
                slug = generate_slug(name)

                # Set default price (you can adjust this)
                base_price = 1000.00

                # Insert product
                cursor.execute("""
                    INSERT INTO products (
                        id, name, slug, description, short_description, sku,
                        base_price, currency, category_id, jewelry_sub_type_id,
                        stone_type_id, ring_style_1_id, ring_style_2_id, ring_style_3_id,
                        ring_style_4_id, ring_style_5_id,
                        is_active, is_featured, in_stock, stock_quantity,
                        meta_title, meta_description,
                        created_at, updated_at
                    ) VALUES (
                        gen_random_uuid(), %s, %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s,
                        NOW(), NOW()
                    ) RETURNING id
                """, (
                    name, slug, description, description[:200] if description else '', sku,
                    base_price, 'GBP', category_id, jewelry_sub_type_id,
                    stone_type_id, ring_style_ids[0], ring_style_ids[1], ring_style_ids[2],
                    ring_style_ids[3], ring_style_ids[4],
                    True, False, True, 1,
                    name, description[:160] if description else ''
                ))

                product_id = cursor.fetchone()['id']

                # Add stone shape relationship if available
                if stone_shape_id:
                    cursor.execute("""
                        INSERT INTO product_stone_shapes (id, product_id, stone_shape_id, created_at, updated_at)
                        VALUES (gen_random_uuid(), %s, %s, NOW(), NOW())
                    """, (product_id, stone_shape_id))

                print(f'  [{row_idx-1}/{engagement_sheet.max_row-1}] IMPORTED: {sku} - {name}')
                imported_count += 1

            except Exception as e:
                print(f'  [{row_idx-1}/{engagement_sheet.max_row-1}] ERROR: {sku} - {name}: {str(e)}')
                error_count += 1
                continue

        # Commit all changes
        conn.commit()

        # Print summary
        print()
        print('='*80)
        print('IMPORT SUMMARY')
        print('='*80)
        print(f'Total rows processed: {engagement_sheet.max_row - 1}')
        print(f'Successfully imported: {imported_count}')
        print(f'Skipped (already exists or invalid): {skipped_count}')
        print(f'Errors: {error_count}')
        print()
        print('Import completed successfully!')

    except Exception as e:
        conn.rollback()
        print(f'\nFATAL ERROR: {str(e)}')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()

if __name__ == '__main__':
    import_products()
